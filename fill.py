# -*- coding: utf-8 -*-
"""
把一筆訂房資料填進對應的 Excel 模板，回傳填好的 Excel（BytesIO）。
每一筆都會產生一個全新的獨立檔案，不會動到原始模板。
"""
import re
import openpyxl
from openpyxl.styles import Alignment, Border, Side
from io import BytesIO
from datetime import datetime, timezone, timedelta
from config import HOTELS, TEMPLATES_DIR, resolve_hotel

# 台灣時區（GMT+8，無日光節約時間）；Render 伺服器為 UTC，
# 填表日期必須用台灣時間，否則會慢一天。
TAIWAN_TZ = timezone(timedelta(hours=8))

# 簡體/繁體 與常見異體字正規化（讓使用者打的簡稱對到模板正式名）
_CHAR_MAP = {
    "槟": "檳", "双": "雙", "牀": "床", "烟": "煙",
    "达": "達", "台": "臺", "伦": "倫", "汇": "匯",
    "门": "門", "个": "個", "东": "東", "厅": "廳",
    "园": "園",
}
# 房型名中的泛詞（去掉後比對核心字，提升簡稱命中率）
# 注意：不要用會把整個名稱拆光的詞（如「客房」「房」「铁塔」），
#       否則核心變成空字串，空字串是任何字串的子串會全部誤命中。
# 也不放「豪華/尊貴」等前綴詞：否則「梅費爾套房」與「豪華梅費爾套房」
#       會被摺成同一核心字而誤命中（交由第三輪 exact-first 處理）。
_GENERIC = [
    "套房", "大床", "雙床", "雙人床", "人", "典雅",
    "景觀", "金光景", "尊貴", "泳池", "有泳池", "-",
]


def _norm(s):
    s = str(s)
    s = re.sub(r"[（）()]", "", s)  # 去掉全/半形括號（房型註記如「天御别墅（四卧室）」）
    for a, b in _CHAR_MAP.items():
        s = s.replace(a, b)
    return re.sub(r"\s+", "", s).lower()


def _core(s):
    s = _norm(s)
    for g in _GENERIC:
        s = s.replace(g, "")
    return s


# ---------------------------------------------------------------------------
# 訂單摘要 Sheet：固定在產出的 Excel 加一張「訂單摘要」，列出
# 代理 / 訂單編號 / 英文姓名 / 中文姓名 / 入住日期 / 退房日期 / 晚數。
# （與原本的客人清單 Sheet1 互不干擾）
# ---------------------------------------------------------------------------
SUMMARY_SHEET = "訂單摘要"
SUMMARY_HEADERS = ["代理", "訂單編號", "英文姓名", "中文姓名", "入住日期", "退房日期"]
_SUMMARY_ALIGN = Alignment(horizontal="center", vertical="center")
_SUMMARY_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _nights(checkin, checkout):
    """由 入住 / 退房 計算晚數；失敗或無效回傳空字串。"""
    try:
        a = datetime.strptime(_norm_date(checkin), "%Y/%m/%d")
        b = datetime.strptime(_norm_date(checkout), "%Y/%m/%d")
        n = (b - a).days
        return n if n > 0 else ""
    except Exception:
        return ""


def _text_width(s):
    """估算字串顯示寬度：中文/全形字約 2 寬，英文/數字約 1 寬。"""
    if s is None or s == "":
        return 0
    return sum(2 if ord(ch) > 127 else 1 for ch in str(s))


def _fill_summary_sheet(wb, booking, guests, primary):
    if SUMMARY_SHEET in wb.sheetnames:
        ws = wb[SUMMARY_SHEET]
    else:
        ws = wb.create_sheet(SUMMARY_SHEET)

    for c, h in enumerate(SUMMARY_HEADERS, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.alignment = _SUMMARY_ALIGN

    sur, fir = split_en_name(primary.get("en_name", ""), primary.get("cn_name", ""))
    en_full = ",".join(x for x in (sur, fir) if x)
    cn_name = primary.get("cn_name", "")
    # 中文姓名在訂單摘要一律顯示繁體
    try:
        from zhconv import convert as _to_trad
        if cn_name:
            cn_name = _to_trad(cn_name, "zh-tw")
    except Exception:
        pass
    row = [
        booking.get("代理", ""),
        booking.get("訂單編號", ""),
        en_full,
        cn_name,
        _norm_date(booking.get("入住", "")),
        _norm_date(booking.get("退房", "")),
    ]
    for c, v in enumerate(row, start=1):
        cell = ws.cell(row=2, column=c, value=v)
        cell.alignment = _SUMMARY_ALIGN

    # 所有標題與資料格加上細框線
    for r in range(1, 3):
        for c in range(1, len(SUMMARY_HEADERS) + 1):
            ws.cell(row=r, column=c).border = _SUMMARY_BORDER

    # 自動調整欄寬：依標題與資料中最寬者；中文約 2 寬、英文/數字約 1 寬，加 2 安全邊距。
    for c, h in enumerate(SUMMARY_HEADERS, start=1):
        col_letter = openpyxl.utils.get_column_letter(c)
        need = max(_text_width(h), _text_width(row[c - 1])) + 2
        ws.column_dimensions[col_letter].width = max(need, 8)


_MONTHS_EN = {"JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6,
               "JUL": 7, "AUG": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12}


def _norm_date(value):
    """把任意常見日期格式歸一為 YYYY/MM/DD。支援：
    - YYYY/M/D、YYYY.M.D、YYYY-M-D
    - YYYY年M月D日
    - YYYYMMDD（8 位純數字）
    - DD/MM/YYYY、DD.MM.YYYY、DD-MMM-YYYY（啟發式：任一 >12 則唯一確定，否則預設 DD/MM）
    - DD MON YYYY（如 18 AUG 1984，大小寫不敏感，可含空白）
    - M月D日、M/D、M.D（無年份）→ 補台灣當年
    無法解析時原樣保留。
    """
    s = re.sub(r"\s+", " ", str(value or "")).strip()
    if not s:
        return ""

    # DD MON YYYY（如 "18 AUG 1984"、"18-Aug-1984"、"18AUG1984"）
    m = re.match(
        r"^(\d{1,2})[ ./\-]*(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)"
        r"[A-Z]*[ ./\-]*(\d{4})$",
        s, re.IGNORECASE,
    )
    if m:
        d, mon, y = m.groups()
        return f"{int(y):04d}/{_MONTHS_EN[mon.upper()[:3]]:02d}/{int(d):02d}"

    # 去空白後比對
    s_nospace = re.sub(r"\s+", "", s)

    # 8 位純數字 YYYYMMDD
    m = re.match(r"^(\d{4})(\d{2})(\d{2})$", s_nospace)
    if m:
        y, mo, d = m.groups()
        return f"{int(y):04d}/{int(mo):02d}/{int(d):02d}"

    # YYYY/M/D、YYYY.M.D、YYYY-M-D
    m = re.match(r"^(\d{4})[./\-](\d{1,2})[./\-](\d{1,2})$", s_nospace)
    if m:
        y, mo, d = m.groups()
        return f"{int(y):04d}/{int(mo):02d}/{int(d):02d}"

    # YYYY年M月D日
    m = re.match(r"^(\d{4})年(\d{1,2})月(\d{1,2})日?$", s_nospace)
    if m:
        y, mo, d = m.groups()
        return f"{int(y):04d}/{int(mo):02d}/{int(d):02d}"

    # DD/MM/YYYY 或 DD.MM.YYYY（啟發式：a>12 必為日；b>12 必為日；都 ≤12 預設 DD/MM）
    m = re.match(r"^(\d{1,2})[./\-](\d{1,2})[./\-](\d{4})$", s_nospace)
    if m:
        a, b, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if a > 12 and b <= 12:
            d, mo = a, b
        elif b > 12 and a <= 12:
            d, mo = b, a
        else:
            d, mo = a, b  # 中文環境預設日/月
        return f"{y:04d}/{mo:02d}/{d:02d}"

    # 中文 M月D日（無年份）
    m = re.match(r"^(\d{1,2})月(\d{1,2})日?$", s_nospace)
    if m:
        mo, d = int(m.group(1)), int(m.group(2))
        return f"{datetime.now(TAIWAN_TZ).year:04d}/{mo:02d}/{d:02d}"

    # M/D 或 M.D（無年份）
    m = re.match(r"^(\d{1,2})[./\-](\d{1,2})$", s_nospace)
    if m:
        mo, d = int(m.group(1)), int(m.group(2))
        return f"{datetime.now(TAIWAN_TZ).year:04d}/{mo:02d}/{d:02d}"

    return s


# 常見中文姓氏 → 護照拼音（大寫）。多字姓放前面，優先匹配。
_SURNAMES_PINYIN = {
    # 複姓
    "歐陽": "OUYANG", "欧阳": "OUYANG", "太史": "TAISHI", "上官": "SHANGGUAN",
    "東方": "DONGFANG", "东方": "DONGFANG", "諸葛": "ZHUGE", "诸葛": "ZHUGE",
    "司馬": "SIMA", "司马": "SIMA", "皇甫": "HUANGFU", "尉遲": "YUCHI",
    "公孙": "GONGSUN", "公孫": "GONGSUN", "慕容": "MURONG", "司徒": "SITU",
    "司空": "SIKONG", "令狐": "LINGHU", "軒轅": "XUANYUAN", "轩辕": "XUANYUAN",
    "南宮": "NANGONG", "南宫": "NANGONG", "夏侯": "XIAHOU", "聞人": "WENREN",
    "鲜于": "XIANYU", "鮮于": "XIANYU", "贺兰": "HELAN", "賀蘭": "HELAN",
    "宇文": "YUWEN", "呼延": "HUYAN", "西門": "XIMEN", "西门": "XIMEN",
    "東郭": "DONGGUO", "东郭": "DONGGUO", "南門": "NANMEN", "南门": "NANMEN",
    "百里": "BAILI", "完顏": "WANYAN", "完颜": "WANYAN", "獨孤": "DUGU", "独孤": "DUGU",
    # 單姓（依頻率與常用度排序，不完整但覆蓋絕大多數酒店訂單）
    "趙": "ZHAO", "赵": "ZHAO", "錢": "QIAN", "钱": "QIAN", "孫": "SUN", "孙": "SUN",
    "李": "LI", "周": "ZHOU", "吳": "WU", "吴": "WU", "鄭": "ZHENG", "郑": "ZHENG",
    "王": "WANG", "馮": "FENG", "冯": "FENG", "陳": "CHEN", "陈": "CHEN", "褚": "CHU",
    "衛": "WEI", "卫": "WEI", "蔣": "JIANG", "蒋": "JIANG", "沈": "SHEN", "韓": "HAN",
    "韩": "HAN", "楊": "YANG", "杨": "YANG", "朱": "ZHU", "秦": "QIN", "尤": "YOU",
    "許": "XU", "许": "XU", "何": "HE", "呂": "LYU", "吕": "LYU", "施": "SHI", "張": "ZHANG",
    "张": "ZHANG", "孔": "KONG", "曹": "CAO", "嚴": "YAN", "严": "YAN", "華": "HUA",
    "华": "HUA", "金": "JIN", "魏": "WEI", "陶": "TAO", "姜": "JIANG", "戚": "QI",
    "謝": "XIE", "谢": "XIE", "鄒": "ZOU", "邹": "ZOU", "喻": "YU", "柏": "BAI",
    "水": "SHUI", "竇": "DOU", "窦": "DOU", "章": "ZHANG", "雲": "YUN", "云": "YUN",
    "蘇": "SU", "苏": "SU", "潘": "PAN", "葛": "GE", "奚": "XI", "范": "FAN",
    "彭": "PENG", "郎": "LANG", "魯": "LU", "鲁": "LU", "韋": "WEI", "韦": "WEI",
    "昌": "CHANG", "馬": "MA", "马": "MA", "苗": "MIAO", "鳳": "FENG", "凤": "FENG",
    "花": "HUA", "方": "FANG", "俞": "YU", "任": "REN", "袁": "YUAN", "柳": "LIU",
    "鮑": "BAO", "鲍": "BAO", "史": "SHI", "唐": "TANG", "費": "FEI", "费": "FEI",
    "廉": "LIAN", "岑": "CEN", "薛": "XUE", "雷": "LEI", "賀": "HE", "贺": "HE",
    "倪": "NI", "湯": "TANG", "汤": "TANG", "滕": "TENG", "殷": "YIN", "羅": "LUO",
    "罗": "LUO", "畢": "BI", "毕": "BI", "郝": "HAO", "鄔": "WU", "邬": "WU",
    "安": "AN", "常": "CHANG", "樂": "LE", "乐": "LE", "于": "YU", "時": "SHI",
    "时": "SHI", "傅": "FU", "付": "FU", "皮": "PI", "卞": "BIAN", "齊": "QI",
    "齐": "QI", "康": "KANG", "伍": "WU", "余": "YU", "元": "YUAN", "卜": "BU",
    "顧": "GU", "顾": "GU", "孟": "MENG", "平": "PING", "黃": "HUANG", "黄": "HUANG",
    "和": "HE", "穆": "MU", "蕭": "XIAO", "萧": "XIAO", "尹": "YIN", "姚": "YAO",
    "邵": "SHAO", "湛": "ZHAN", "汪": "WANG", "祁": "QI", "毛": "MAO", "禹": "YU",
    "狄": "DI", "米": "MI", "貝": "BEI", "贝": "BEI", "明": "MING", "臧": "ZANG",
    "計": "JI", "计": "JI", "伏": "FU", "成": "CHENG", "戴": "DAI", "談": "TAN",
    "谈": "TAN", "宋": "SONG", "茅": "MAO", "龐": "PANG", "庞": "PANG", "熊": "XIONG",
    "紀": "JI", "纪": "JI", "舒": "SHU", "屈": "QU", "項": "XIANG", "项": "XIANG",
    "祝": "ZHU", "董": "DONG", "梁": "LIANG", "杜": "DU", "阮": "RUAN", "藍": "LAN",
    "蓝": "LAN", "閔": "MIN", "闵": "MIN", "席": "XI", "季": "JI", "麻": "MA",
    "強": "QIANG", "强": "QIANG", "賈": "JIA", "贾": "JIA", "路": "LU", "婁": "LOU",
    "娄": "LOU", "危": "WEI", "江": "JIANG", "童": "TONG", "顏": "YAN", "颜": "YAN",
    "郭": "GUO", "梅": "MEI", "盛": "SHENG", "林": "LIN", "刁": "DIAO", "鍾": "ZHONG",
    "钟": "ZHONG", "徐": "XU", "邱": "QIU", "駱": "LUO", "骆": "LUO", "高": "GAO",
    "夏": "XIA", "蔡": "CAI", "田": "TIAN", "樊": "FAN", "胡": "HU", "凌": "LING",
    "霍": "HUO", "虞": "YU", "萬": "WAN", "万": "WAN", "支": "ZHI", "柯": "KE",
    "管": "GUAN", "盧": "LU", "卢": "LU", "莫": "MO", "房": "FANG", "裘": "QIU",
    "繆": "MIAO", "缪": "MIAO", "干": "GAN", "應": "YING", "应": "YING", "宗": "ZONG",
    "丁": "DING", "宣": "XUAN", "賁": "BEN", "贲": "BEN", "鄧": "DENG", "邓": "DENG",
    "郁": "YU", "單": "SHAN", "单": "SHAN", "杭": "HANG", "洪": "HONG", "包": "BAO",
    "石": "SHI", "崔": "CUI", "吉": "JI", "鈕": "NIU", "钮": "NIU", "龔": "GONG",
    "龚": "GONG", "程": "CHENG", "嵇": "JI", "邢": "XING", "裴": "PEI", "陸": "LU",
    "陆": "LU", "榮": "RONG", "荣": "RONG", "翁": "WENG", "荀": "XUN", "羊": "YANG",
    "惠": "HUI", "甄": "ZHEN", "麴": "QU", "封": "FENG", "芮": "RUI", "羿": "YI",
    "儲": "CHU", "储": "CHU", "靳": "JIN", "汲": "JI", "邴": "BING", "糜": "MI",
    "松": "SONG", "井": "JING", "段": "DUAN", "富": "FU", "巫": "WU", "烏": "WU",
    "乌": "WU", "焦": "JIAO", "巴": "BA", "弓": "GONG", "牧": "MU", "山": "SHAN",
    "谷": "GU", "車": "CHE", "车": "CHE", "侯": "HOU", "蓬": "PENG", "全": "QUAN",
    "郗": "XI", "班": "BAN", "仰": "YANG", "秋": "QIU", "仲": "ZHONG", "伊": "YI",
    "宮": "GONG", "宫": "GONG", "寧": "NING", "宁": "NING", "仇": "QIU", "欒": "LUAN",
    "栾": "LUAN", "暴": "BAO", "甘": "GAN", "鈄": "TOU", "钭": "TOU", "厲": "LI",
    "厉": "LI", "戎": "RONG", "祖": "ZU", "武": "WU", "符": "FU", "劉": "LIU",
    "刘": "LIU", "景": "JING", "詹": "ZHAN", "束": "SHU", "龍": "LONG", "龙": "LONG",
    "葉": "YE", "叶": "YE", "司": "SI", "韶": "SHAO", "郜": "GAO", "黎": "LI",
    "薊": "JI", "蓟": "JI", "薄": "BO", "印": "YIN", "宿": "SU", "白": "BAI",
    "懷": "HUAI", "怀": "HUAI", "蒲": "PU", "邰": "TAI", "鄂": "E", "索": "SUO",
    "咸": "XIAN", "籍": "JI", "賴": "LAI", "赖": "LAI", "卓": "ZHUO", "藺": "LIN",
    "蔺": "LIN", "屠": "TU", "蒙": "MENG", "池": "CHI", "喬": "QIAO", "乔": "QIAO",
    "陰": "YIN", "阴": "YIN", "鬱": "YU", "郁": "YU", "胥": "XU", "能": "NENG",
    "蒼": "CANG", "苍": "CANG", "雙": "SHUANG", "双": "SHUANG", "聞": "WEN", "闻": "WEN",
    "莘": "SHEN", "黨": "DANG", "党": "DANG", "翟": "ZHAI", "譚": "TAN", "谭": "TAN",
    "貢": "GONG", "贡": "GONG", "勞": "LAO", "劳": "LAO", "逄": "PANG", "姬": "JI",
    "申": "SHEN", "扶": "FU", "堵": "DU", "冉": "RAN", "宰": "ZAI", "酈": "LI",
    "郦": "LI", "雍": "YONG", "郤": "XI", "璩": "QU", "桑": "SANG", "桂": "GUI",
    "濮": "PU", "牛": "NIU", "壽": "SHOU", "寿": "SHOU", "通": "TONG", "邊": "BIAN",
    "边": "BIAN", "扈": "HU", "燕": "YAN", "冀": "JI", "郟": "JIA", "郏": "JIA",
    "浦": "PU", "尚": "SHANG", "農": "NONG", "农": "NONG", "溫": "WEN", "温": "WEN",
    "別": "BIE", "别": "BIE", "莊": "ZHUANG", "庄": "ZHUANG", "晏": "YAN", "柴": "CHAI",
    "瞿": "QU", "閻": "YAN", "阎": "YAN", "充": "CHONG", "慕": "MU", "連": "LIAN",
    "连": "LIAN", "茹": "RU", "習": "XI", "习": "XI", "宦": "HUAN", "艾": "AI",
    "魚": "YU", "鱼": "YU", "容": "RONG", "古": "GU", "易": "YI", "慎": "SHEN",
    "戈": "GE", "廖": "LIAO", "庾": "YU", "終": "ZHONG", "终": "ZHONG", "暨": "JI",
    "居": "JU", "衡": "HENG", "步": "BU", "都": "DU", "耿": "GENG", "滿": "MAN",
    "满": "MAN", "弘": "HONG", "匡": "KUANG", "國": "GUO", "国": "GUO", "文": "WEN",
    "寇": "KOU", "廣": "GUANG", "广": "GUANG", "祿": "LU", "禄": "LU", "闕": "QUE",
    "阙": "QUE", "東": "DONG", "东": "DONG", "歐": "OU", "欧": "OU", "殳": "SHU",
    "沃": "WO", "利": "LI", "越": "YUE", "隆": "LONG", "師": "SHI", "师": "SHI",
    "鞏": "GONG", "巩": "GONG", "聶": "NIE", "聂": "NIE", "晁": "CHAO", "勾": "GOU",
    "敖": "AO", "融": "RONG", "冷": "LENG", "訾": "ZI", "辛": "XIN", "闞": "KAN",
    "阚": "KAN", "那": "NA", "簡": "JIAN", "简": "JIAN", "饒": "RAO", "饶": "RAO",
    "空": "KONG", "曾": "ZENG", "毋": "WU", "沙": "SHA", "乜": "NIE", "養": "YANG",
    "养": "YANG", "鞠": "JU", "須": "XU", "须": "XU", "豐": "FENG", "丰": "FENG",
    "巢": "CHAO", "關": "GUAN", "关": "GUAN", "蒯": "KUAI", "相": "XIANG", "查": "ZHA",
    "後": "HOU", "后": "HOU", "荊": "JING", "荆": "JING", "紅": "HONG", "红": "HONG",
    "游": "YOU", "竺": "ZHU", "權": "QUAN", "权": "QUAN", "逯": "LU", "蓋": "GE",
    "盖": "GE", "益": "YI", "桓": "HUAN", "公": "GONG", "牟": "MOU", "哈": "HA",
    "言": "YAN", "福": "FU", "肖": "XIAO", "區": "OU", "区": "OU", "麥": "MAI",
    "麦": "MAI", "佟": "TONG", "靖": "JING", "湛": "ZHAN", "謀": "MOU", "谋": "MOU",
    "譚": "TAN", "谭": "TAN", "尋": "XUN", "寻": "XUN", "棘": "JI", "銀": "YIN",
    "银": "YIN", "佴": "ER", "伯": "BO", "賞": "SHANG", "赏": "SHANG", "松": "SONG",
    "段": "DUAN", "甄": "ZHEN", "尉": "WEI", "遲": "CHI", "迟": "CHI", "公": "GONG",
    "孫": "SUN", "孙": "SUN", "公": "GONG", "冶": "YE", "淳": "CHUN", "於": "YU",
    "乜": "NIE", "督": "DU", "仉": "ZHANG", "司": "SI", "邛": "QIONG", "僪": "YU",
    "都": "DU", "粲": "CAN", "僧": "SENG", "薩": "SA", "萨": "SA", "隗": "KUI",
    "穰": "RANG", "還": "HUAN", "邴": "BING", "雒": "LUO", "臧": "ZANG", "紅": "HONG",
    "红": "HONG",
}


def split_en_name(en: str, cn_name: str = None):
    """把英文姓名拆成 (姓, 名)，並一律輸出大寫。
    不論輸入是半形逗號、全形逗號、斜線、點號或空白分隔，都會正確拆開。
    若英文無分隔且提供中文姓名，會用中文姓的拼音自動切開（如 ZHANGYITING + 张依婷 -> ZHANG,YITING）。
    範例：'QU,SHENZHONG'   -> ('QU', 'SHENZHONG')
          'TANG/QINGPING'  -> ('TANG', 'QINGPING')
          'ZHOU.YINHUI'    -> ('ZHOU', 'YINHUI')
          'SHEN DAN'       -> ('SHEN', 'DAN')
          'QIU，JIELEI'    -> ('QIU', 'JIELEI')  （全形逗號）
          'qiu,jielei'     -> ('QIU', 'JIELEI')  （強制大寫）
          'ZHANGYITING'    -> ('ZHANG', 'YITING')  （無分隔，按中文姓切）
    """
    en = (en or "").strip()
    if not en:
        return "", ""
    # 全形逗號先統一成半形逗號
    en = en.replace("，", ",")
    # 逗號 / 斜線 / 點號分隔：前=姓 後=名
    for sep in (",", "/", "."):
        if sep in en:
            a, b = en.split(sep, 1)
            return a.strip().upper(), b.strip().upper()
    # 純空白分隔：中文習慣第一個字是姓
    toks = en.split()
    if len(toks) >= 2:
        return toks[0].upper(), " ".join(toks[1:]).upper()
    # 無分隔：嘗試用中文姓的拼音切開
    if cn_name:
        cn = str(cn_name).strip()
        # 先試複姓（2 字），再試單姓（1 字）
        for n in (2, 1):
            if len(cn) >= n:
                py = _SURNAMES_PINYIN.get(cn[:n])
                if py and en.upper().startswith(py):
                    return py, en[len(py):].upper()
    return "", en.upper()


def _set_merged_cell(ws, coord, value):
    """安全寫入儲存格；若 coord 位於合併儲存格內，則改寫該合併範圍左上角那一格。"""
    r, c = openpyxl.utils.coordinate_to_tuple(coord)
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= r <= rng.max_row and rng.min_col <= c <= rng.max_col:
            coord = f"{openpyxl.utils.get_column_letter(rng.min_col)}{rng.min_row}"
            break
    ws[coord] = value


def _write_labeled(ws, coord, value):
    """寫入數值；若該格原本就含標籤文字（如康萊德「人數 Pax(位):」、
    「房數 No.of Rooms (間):」標籤與填空合一），保留標籤並接上數值；
    其他飯店此格為空白填空格，直接寫入數值。"""
    existing = ws[coord].value
    if existing is not None and str(existing).strip():
        txt = str(existing).strip()
        if re.search(r"\d\s*$", txt):
            ws[coord] = txt          # 已含數值就不重複加
        else:
            ws[coord] = f"{txt}{value}"
    else:
        ws[coord] = value


# ---------------------------------------------------------------------------
# 房型勾選：匹配 booking 房型 → 在模板對應儲存格打 ✓
# ---------------------------------------------------------------------------
def _match_room_type(room_types: list, room_type_str: str):
    """在飯店房型列表中尋找匹配項，回傳儲存格座標或 None。
    matching priority: exact code → code group → exact Chinese → weighted ngram
    """
    if not room_type_str:
        return None
    s = (room_type_str or "").strip()
    su = s.upper()
    s_clean = re.sub(r"[（()].*", "", s).strip()

    # Pass 1: exact matches (code / name)
    for cell, code, cn in room_types:
        if su == code.upper():
            return cell
        for c in code.split("/"):
            if su == c.strip().upper():
                return cell
        if s_clean == cn:
            return cell

    # Pass 2: weighted scoring (core match + substring + ngram)
    sc = _core(s_clean)
    best_cell, best_score, best_ngram_hits, best_len = None, 0, 0, 9999
    for cell, code, cn in room_types:
        score = 0
        # Core match (stripped of generic terms like 套房/大床/雙床)
        cnc = _core(cn)
        if sc and cnc:
            if sc in cnc or cnc in sc:
                score += 50
        # Substring containment
        if s_clean in cn or cn in s_clean:
            score += 100
        # N-gram intersection: longer shared substrings → higher weight
        ngram_hits = 0
        for size in (4, 3, 2):
            cn_set = {cn[i:i+size] for i in range(len(cn)-size+1)}
            s_set = {s_clean[i:i+size] for i in range(len(s_clean)-size+1)}
            hits = len(cn_set & s_set)
            score += hits * size
            ngram_hits += hits
        # tiebreaker: 同分 → ngram 命中多者優先 → 名字短者優先
        if score == best_score:
            if ngram_hits > best_ngram_hits:
                best_cell, best_score, best_ngram_hits = cell, score, ngram_hits
            elif ngram_hits == best_ngram_hits and len(cn) < best_len:
                best_cell, best_score, best_ngram_hits, best_len = cell, score, ngram_hits, len(cn)
        elif score > best_score:
            best_cell, best_score, best_ngram_hits, best_len = cell, score, ngram_hits, len(cn)

    return best_cell if best_score > 0 else None


def _check_room_cell(ws, cell_coord: str):
    """在房型格填入 ✓。若該格原本含 (  ) 格式 → 取代為 (✓)；
    否則直接寫入 ✓ 在原值之後（保留房型名稱文字）。"""
    existing = ws[cell_coord].value or ""
    txt = str(existing)
    # 範本常見格式：(  ) 房型名 → 改為 (✓) 房型名
    replaced = re.sub(r"\(\s*\)", "(✓)", txt, count=1)
    if replaced != txt:
        ws[cell_coord] = replaced
    else:
        # 無 (  ) 格式就直接接 ✓
        ws[cell_coord] = (txt.strip() + " ✓").strip()


def _check_billing_cell(ws, cell_coord: str):
    """在『收費 Billing』勾選格填入 V：將 (    ) 改為 ( V )。"""
    existing = ws[cell_coord].value or ""
    txt = str(existing)
    replaced = re.sub(r"\(\s*\)", "( V )", txt, count=1)
    if replaced != txt:
        ws[cell_coord] = replaced
    else:
        ws[cell_coord] = (txt.strip() + " V").strip()


def fill_booking(booking: dict) -> BytesIO:
    """
    booking 結構：
    {
      "飯店": "名匯",
      "入住": "2026/07/20",
      "退房": "2026/07/22",
      "房型": "RK" 或 "豪華大床房",
      "件數": "2",            # 房數
      "備注": "高樓層",
      "是否吸煙": "不吸煙",
      "guests": [
         {"cn_name":"渠慎重", "en_name":"QU,SHENZHONG", "dob":"1961/06/11", "idno":"M41646681"},
         ...
      ]
    }
    """
    hotel_key = resolve_hotel(booking.get("飯店", ""))
    if not hotel_key:
        raise ValueError(f"找不到對應飯店：{booking.get('飯店')}")
    cfg = HOTELS[hotel_key]

    wb = openpyxl.load_workbook(TEMPLATES_DIR / cfg["file"])
    mws = wb[cfg["main_sheet"]]
    mc = cfg["main_cells"]
    guests = booking.get("guests", []) or []
    primary = guests[0] if guests else {}

    sur, fir = split_en_name(primary.get("en_name", ""), primary.get("cn_name", ""))
    mws[mc["surname"]] = sur
    mws[mc["firstname"]] = fir
    mws[mc["idno"]] = primary.get("idno", "")
    mws[mc["dob"]] = _norm_date(primary.get("dob", ""))
    mws[mc["checkin"]] = _norm_date(booking.get("入住", ""))
    mws[mc["checkout"]] = _norm_date(booking.get("退房", ""))
    # 房數 / 人數：若該格原本含標籤文字（如康萊德「人數 Pax(位):」、
    # 「房數 No.of Rooms (間):」標籤與填空合一），保留標籤並接上數值；
    # 其他飯店此格為空白填空格，直接寫入數值。
    _write_labeled(mws, mc["rooms"], booking.get("件數", ""))
    _write_labeled(mws, mc["pax"], len(guests))

    # ---- 房型打勾：匹配 booking 房型 → 在對應的 (  ) 格填入 ✓ ----
    room_type_str = (booking.get("房型", "") or "").strip()
    if room_type_str and cfg.get("room_types"):
        cell = _match_room_type(cfg["room_types"], room_type_str)
        if cell:
            _check_room_cell(mws, cell)

    # ---- 收費 Billing 自動打 V：只要產檔就勾選 ----
    if "billing" in mc:
        _check_billing_cell(mws, mc["billing"])

    # 填表日期（右下角 Date: 後面的底線格；只填日期值，保留模板藍色字體格式）
    if "date" in mc:
        _set_merged_cell(mws, mc["date"], datetime.now(TAIWAN_TZ).strftime('%Y/%m/%d'))

    # 備注 + 吸煙
    remark = (booking.get("備注", "") or "").strip()
    smoking = (booking.get("是否吸煙", "") or "").strip()
    # 過濾無意義備注（單純「無」不顯示）
    if remark in ("無", "无", "沒有", "没有", "-"):
        remark = ""
    if smoking:
        # 所有飯店都把吸煙資訊併入備注
        remark = (remark + f"；{smoking}").strip("；")

    # 寫入備注：若該格原本就含欄位標籤（如「特別要求 Special request :」），
    # 保留標籤並把實際備注接在後面，避免覆蓋掉欄位名稱。
    existing = mws[mc["remark"]].value
    if existing and str(existing).strip():
        mws[mc["remark"]] = f"{existing}{remark}" if remark else existing
    else:
        mws[mc["remark"]] = remark

    # ---- 客人清單（部分飯店無獨立客人清單頁，例如康萊德，直接跳過）----
    if cfg.get("guest_sheet") and cfg.get("guest_cols"):
        gws = wb[cfg["guest_sheet"]]
        gc = cfg["guest_cols"]
        first = cfg["guest_first_row"]
        maxr = gws.max_row

        # 清空舊的範例資料
        for r in range(first, maxr + 1):
            for col in gc.values():
                gws[f"{col}{r}"].value = None

        # 客人數超過現有列數 -> 往下加列
        existing = maxr - first + 1
        needed = max(len(guests), 1)
        if needed > existing:
            gws.insert_rows(maxr + 1, needed - existing)

        for i, g in enumerate(guests):
            r = first + i
            if "cn_name" in gc:
                gws[f"{gc['cn_name']}{r}"] = g.get("cn_name", "")
            if "en_name" in gc:
                gws[f"{gc['en_name']}{r}"] = g.get("en_name", "")
            if "dob" in gc:
                gws[f"{gc['dob']}{r}"] = g.get("dob", "")
            if "idno" in gc:
                gws[f"{gc['idno']}{r}"] = g.get("idno", "")
            if "roomtype" in gc:
                gws[f"{gc['roomtype']}{r}"] = booking.get("房型", "")
            if "smoking" in gc:
                gws[f"{gc['smoking']}{r}"] = smoking

    # ---- 訂單摘要 Sheet（固定新增，與客人清單並存）----
    _fill_summary_sheet(wb, booking, guests, primary)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def _short_date(raw: str) -> str:
    """把日期轉成 月.日 格式（去掉年份），如 2026/07/20 → 07.20。
    解析失敗或非預期長度則回傳空字串。"""
    d = _norm_date(raw)
    if not d:
        return ""
    parts = d.split("/")
    if len(parts) == 3:
        return f"{parts[1]}.{parts[2]}"
    return ""


def output_filename(booking: dict) -> str:
    """產出檔名：{入住月.日}-{退房月.日}_{飯店}_{姓名}.xlsx
    年份不顯示；姓名會轉繁體（港澳用字，如 李彦超→李彥超）。
    若缺日期則回退為「訂房」前綴，避免多出分隔符。"""
    hotel = resolve_hotel(booking.get("飯店", "")) or "訂房"
    g0 = (booking.get("guests") or [{}])[0]
    name = g0.get("cn_name") or g0.get("en_name") or ""
    # 姓名轉繁體
    try:
        from zhconv import convert as _to_trad
        if name:
            name = _to_trad(name, "zh-tw")
    except Exception:
        pass
    ci = _short_date(booking.get("入住", ""))
    co = _short_date(booking.get("退房", ""))
    if ci and co:
        prefix = f"{ci}-{co}"
    elif ci:
        prefix = ci
    elif co:
        prefix = co
    else:
        prefix = "訂房"
    return f"{prefix}_{hotel}_{name}.xlsx"


# ---------------------------------------------------------------------------
# 中文 / 英文姓名拼音自動核對
# ---------------------------------------------------------------------------
_NAME_TITLES = {
    "MR", "MRS", "MS", "MISS", "DR", "MASTER", "MSTR", "MDM", "MADAM",
    "SIR", "PROF",
}

# 台灣/香港常見羅馬拼音對照（大陸拼音 → 各地變體）
# 用於中文姓名核對：只要符合大陸拼音 或 台灣/香港拼音其一，就不提示。
_SURNAME_TW_VARIANTS = {
    # 大陸拼音 → {台灣/香港變體}
    "LI": {"LEE"},          # 李
    "WANG": {"WONG"},       # 王（香港）
    "ZHANG": {"CHANG"},     # 張
    "ZHENG": {"CHENG"},     # 鄭
    "ZHOU": {"CHOU"},       # 周
    "ZHU": {"CHU"},         # 朱
    "ZHAO": {"CHAO"},       # 趙
    "XU": {"HSU"},          # 徐 / 許
    "XIAO": {"HSIAO"},      # 蕭
    "XIE": {"HSIEH"},       # 謝
    "QIU": {"CHIU"},        # 邱
    "QIAN": {"CHIEN"},      # 錢
    "QIN": {"CHIN"},        # 秦
    "GUO": {"KUO"},         # 郭
    "CAI": {"TSAI"},        # 蔡
    "ZENG": {"TSENG"},      # 曾
    "JIANG": {"CHIANG"},    # 蔣 / 江
    "JIA": {"CHIA"},        # 賈
    "CHEN": {"CHAN"},        # 陳（香港）
    "LIU": {"LAU"},         # 劉（香港）
    "HUANG": {"WONG"},      # 黃（香港）
    "HE": {"HO"},           # 何（香港）
    "LUO": {"LO"},          # 羅（香港）
    "WU": {"NG"},           # 吳（香港）
    "LIN": {"LAM"},         # 林（香港）
    "YANG": {"YEUNG"},      # 楊（香港）
}

# Wade-Giles 聲母對應表（用於名字部分的大陸→台灣轉換）
_WG_INITIAL_MAP = [
    ("ZH", "CH"), ("CH", "CH"), ("SH", "SH"),
    ("X", "HS"),   ("Q", "CH"),
    ("C", "TS"),   ("Z", "TS"),
    ("J", "CH"),   ("G", "K"),
]


def _to_wade_giles(py: str) -> set:
    """把一個音節的大陸拼音轉成所有可能的 Wade-Giles 變體。"""
    py = py.upper()
    out = {py}
    for mp, wg in _WG_INITIAL_MAP:
        if py.startswith(mp):
            wg_var = wg + py[len(mp):]
            out.add(wg_var)
            break  # 只換第一組匹配的聲母
    return out


def _cn_to_pinyin(cn: str) -> str:
    """把中文姓名轉成無聲調、無空白的大寫拼音（如 张依婷 -> ZHANGYITING）。
    姓氏優先用 _SURNAMES_PINYIN 表（處理多音姓），其餘用 pypinyin。"""
    from pypinyin import lazy_pinyin
    cn = str(cn or "").strip()
    for n in (2, 1):
        if len(cn) >= n and cn[:n] in _SURNAMES_PINYIN:
            sur = _SURNAMES_PINYIN[cn[:n]]
            rest = "".join(lazy_pinyin(cn[n:])).upper()
            return (sur + rest)
    return "".join(lazy_pinyin(cn)).upper()


def _cn_pinyin_acceptable(cn: str):
    """回傳中文姓名可接受的英文全拼集合（大寫、無分隔）。
    涵蓋：姓+名、名+姓、只給姓；多音字所有讀音組合；
    **台灣/香港拼音變體**（如 李=LI/LEE、張=ZHANG/CHANG）。
    無法轉拼音時回傳 None。
    """
    from pypinyin import pinyin, Style
    cn = str(cn or "").strip()
    if not cn:
        return None
    # 決定姓（複姓優先），姓拼音優先取權威表
    sur_len, sur_set = 0, set()
    for n in (2, 1):
        if len(cn) >= n and cn[:n] in _SURNAMES_PINYIN:
            sur_py = _SURNAMES_PINYIN[cn[:n]]
            sur_set = {sur_py}
            # 加台灣/香港變體
            sur_set |= _SURNAME_TW_VARIANTS.get(sur_py, set())
            sur_len = n
            break
    if not sur_set:
        first = pinyin(cn[:1], style=Style.NORMAL, heteronym=True)[0]
        sur_set = {o.upper() for o in first}
        # 對每種讀音加入 Wade-Giles 變體（無權威表時）
        for o in list(sur_set):
            sur_set |= _to_wade_giles(o)
        sur_len = 1
    rest = cn[sur_len:]
    # 名（可能多字、多音）→ 笛卡兒積展開所有讀音組合，並加入 Wade-Giles 變體
    given_set = {""}
    if rest:
        per = pinyin(rest, style=Style.NORMAL, heteronym=True)
        for options in per:
            expanded = set()
            for o in options:
                expanded |= _to_wade_giles(o.upper())
            given_set = {g + o for g in given_set for o in expanded}

    def _clean(s):
        return re.sub(r"[^A-Z]", "", s)  # 去 ü 等非 A-Z 字元

    acceptable = set()
    for s in sur_set:
        s = _clean(s)
        acceptable.add(s)
        for g in given_set:
            g = _clean(g)
            acceptable.add(s + g)
            acceptable.add(g + s)
    acceptable.discard("")
    return acceptable


def verify_name_match(cn_name: str, en_name: str):
    """比對中文姓名與英文姓名拼音是否一致（嚴格音節比對）。
    回傳 (ok, message)：ok=True 表示相符或無法判斷；ok=False 附上提示訊息。
    判斷規則：
      - 任一方為空、或無法轉拼音 -> 視為 ok（不誤報）
      - 英文（去稱謂 MR/MS…、去分隔符）需完全等於下列其一才算相符：
          · 姓+名、名+姓（多音字所有讀音組合皆納入）
          · 只給姓（容許只填姓氏）
      - 否則視為不符，回傳提示（附預期拼音）
    """
    cn = (cn_name or "").strip()
    en = (en_name or "").strip()
    if not cn or not en:
        return True, None
    # 英文正規化：切 token → 去稱謂 → 只留 A-Z
    tokens = [t for t in re.split(r"[,\s/\.．，、]+", en.upper()) if t]
    tokens = [t for t in tokens if t not in _NAME_TITLES]
    en_norm = re.sub(r"[^A-Z]", "", "".join(tokens))
    if not en_norm:
        return True, None
    try:
        acceptable = _cn_pinyin_acceptable(cn)
    except Exception:
        return True, None
    if not acceptable:
        return True, None
    if en_norm in acceptable:
        return True, None

    try:
        expect = _cn_to_pinyin(cn)
    except Exception:
        expect = ""
    tail = f"（大陸拼音：{expect}；台灣/香港如 LEE/CHANG/HSU 也接受）" if expect else ""
    return False, (
        f"⚠️ 中文姓名「{cn}」與英文姓名拼音「{en}」似乎不符，"
        f"請確認英文拼音是否正確{tail}。"
    )


def verify_booking_names(booking: dict):
    """核對訂房資料中所有入住者的 中文/英文 姓名拼音。
    回傳 (all_ok, warnings) ；warnings 為不符的提示清單。
    """
    guests = booking.get("guests") or []
    warnings = []
    for i, g in enumerate(guests, 1):
        ok, msg = verify_name_match(g.get("cn_name", ""), g.get("en_name", ""))
        if not ok and msg:
            prefix = f"第{i}位入住者：" if len(guests) > 1 else ""
            warnings.append(prefix + msg)
    return (len(warnings) == 0, warnings)


# ---------------------------------------------------------------------------
# 年齡檢查：未滿 21 歲提示
# ---------------------------------------------------------------------------
def _compute_age(dob_str: str, checkin_str: str):
    """計算從出生日到入住日的實足年齡；任一日期無法解析時回傳 None。"""
    try:
        dob = datetime.strptime(_norm_date(dob_str), "%Y/%m/%d")
    except Exception:
        return None
    ci = _norm_date(checkin_str) if checkin_str else ""
    if not ci:
        return None
    try:
        ci_date = datetime.strptime(ci, "%Y/%m/%d")
    except Exception:
        return None
    age = ci_date.year - dob.year
    if (ci_date.month, ci_date.day) < (dob.month, dob.day):
        age -= 1
    return age


def verify_guests_age(booking: dict):
    """檢查入住者是否未滿 21 歲。
    回傳 (all_ok, warnings)；warnings 為年齡不足的提示清單。
    """
    guests = booking.get("guests") or []
    checkin = booking.get("入住", "")
    warnings = []
    for i, g in enumerate(guests, 1):
        dob = g.get("dob", "")
        if not dob:
            continue
        age = _compute_age(dob, checkin)
        if age is not None and age < 21:
            cn = g.get("cn_name", "") or g.get("en_name", "")
            name = cn or f"第{i}位"
            prefix = f"第{i}位入住者「{name}」：" if len(guests) > 1 else f"入住者「{name}」："
            warnings.append(
                f"⚠️ {prefix}年齡未滿 21 歲（入住日實足 {age} 歲），"
                f"請確認是否符合飯店入住規定。"
            )
    return (len(warnings) == 0, warnings)

