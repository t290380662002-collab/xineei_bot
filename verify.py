# -*- coding: utf-8 -*-
import openpyxl, glob, os

files = sorted(glob.glob("C:/Users/t2903/WorkBuddy/2026-07-13-01-18-38/output/*.xlsx"))
for f in files:
    print("=" * 78)
    print("FILE:", os.path.basename(f))
    wb = openpyxl.load_workbook(f)
    print("Sheets:", wb.sheetnames)
    ws = wb[wb.sheetnames[0]]
    print("--- 主表單關鍵欄位 ---")
    for coord in ["B16","H16","H17","B18","B20","M20","B21","M21","L14",
                  "C23","G23","C24","G24","C25","G25","C26","G26","C27","G27","C28","G28"]:
        v = ws[coord].value
        if v not in (None, ""):
            print(f"  {coord}: {repr(v)}")
    # 找被打勾的房型
    print("  打勾的房型：", [c for c in ["C23","G23","K23","N23","C24","G24","K24","N24","C25","G25","K25","N25","C26","G26","K26","N26","C27","G27","K27","N27","C28","G28"] if ws[c].value and "(✓)" in str(ws[c].value)])
    # 客人清單
    gname = wb.sheetnames[-1]
    gws = wb[gname]
    print(f"--- 客人清單 ({gname}) ---")
    for row in gws.iter_rows():
        cells = [(c.coordinate, c.value) for c in row if c.value not in (None, "")]
        if cells:
            print("  ", cells)
