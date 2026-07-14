import sys, os
sys.path.insert(0, r"C:/Users/t2903/WorkBuddy/2026-07-13-01-18-38")
from parse_text import parse_booking_text, looks_like_booking
from config import resolve_hotel
from fill import fill_booking
import openpyxl
from io import BytesIO

sample = """入住：7月15日
退房：7月16日
飯店：伦敦人名汇酒店
房型:  伊丽莎白套房  双床
件數：1间1晚
備注：高楼层、无烟房
是否吸煙：无烟房

入住者中文：孙碧春
入住者英文：SUN, BICHUN
出生年月日：2000.05.25
證件號碼：CK6005178
SS 不會讀取"""

print("looks_like_booking:", looks_like_booking(sample))
b = parse_booking_text(sample)
for k,v in b.items():
    print(f"  {k}: {v}")

print("\nresolve_hotel(飯店):", resolve_hotel(b["飯店"]))

bio = fill_booking(b)
wb = openpyxl.load_workbook(BytesIO(bio.getvalue()), data_only=False)
print("sheets:", wb.sheetnames)
ws = wb[wb.sheetnames[0]]
for row in ws.iter_rows():
    for cell in row:
        v = cell.value
        if v and isinstance(v, str) and ("吸煙" in v or "特別要求" in v or "SS" in v or "未對應" in v):
            print(f"  {cell.coordinate}: {v}")
# save
out = r"C:/Users/t2903/WorkBuddy/2026-07-13-01-18-38/output/訂房_TEST_孙碧春.xlsx"
with open(out,"wb") as f:
    f.write(bio.getvalue())
print("saved:", out)
