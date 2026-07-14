import openpyxl
files = {
    "名匯": "output/訂房_名匯_渠慎重.xlsx",
    "威尼斯": "output/訂房_威尼斯_褚國華.xlsx",
    "康萊德": "output/訂房_康萊德_孙碧春.xlsx",
    "御園": "output/訂房_御園_渠慎重.xlsx",
}
for hk, f in files.items():
    wb = openpyxl.load_workbook(f)
    ws = wb.active
    # 找備注與一個房型格
    print(f"[{hk}] sheets={wb.sheetnames} active={ws.title}")
    # 列印 L14 備注
    print(f"   L14(remark)={ws['L14'].value!r}")
