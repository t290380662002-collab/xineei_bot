import openpyxl, glob, os

files = sorted(glob.glob("C:/Users/t2903/WorkBuddy/2026-07-13-01-18-38/templates/*.xlsx"))
for f in files:
    print("=" * 80)
    print("FILE:", os.path.basename(f))
    wb = openpyxl.load_workbook(f)
    print("Sheets:", wb.sheetnames)
    for ws in wb.worksheets:
        print(f"--- Sheet: {ws.title}  dims={ws.dimensions} max_row={ws.max_row} max_col={ws.max_column}")
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and str(cell.value).strip() != "":
                    print(f"  {cell.coordinate}: {repr(cell.value)}")
        print("  Merged:", list(ws.merged_cells.ranges))
